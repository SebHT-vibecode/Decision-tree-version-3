import openpyxl

wb = openpyxl.load_workbook('Data/IQVIA CD Device Catalog_01-Apr-2026_Novo.Nordisk.xlsx')
ws = wb.active

print('Columns:')
for col in range(1, ws.max_column + 1):
    print(ws.cell(1, col).value)

print('\nFirst 5 rows:')
for row in range(1, 6):
    row_data = [ws.cell(row, col).value for col in range(1, ws.max_column + 1)]
    print(row_data)